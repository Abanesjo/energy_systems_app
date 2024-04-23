import numpy as np
import pandas as pd
from pyXSteam.XSteam import XSteam
import fluids
import sigfig as sf
from openpyxl import load_workbook

class node:
    def __init__(self, name="", p=0.0, t=0.0, q=0.0, h=0.0):
        self.name = name
        self.p = p
        self.t = t
        self.q = q
        self.h = h
    def get_parameters(self):
        print(f"{self.name} parameters:\n {self.p:.2f} bara | {self.t:.2f} C\n {self.q:.2f} kg/hr | {self.h:.2f} kJ/kg")
        return [self.p, self.t, self.q, self.h]
    
def pressure_drop(p_avg, t_avg, q, d, epsilon, fitting_resistance, L_p, static_head_m, misc_pressure_drop_m):
    st = XSteam(XSteam.UNIT_SYSTEM_MKS)
    g = 9.81
    D = d / 1000 #d in meters
    V = st.v_pt(p_avg, t_avg)
    my = st.my_pt(p_avg, t_avg)
    vel = 353.678 * q * V / (d**2)
    Re = (vel * D) / (V * my)
    f = fluids.friction_factor(Re, epsilon/D)
    turbulence = "Turbulent" if Re > 2300 else "Laminar"

    # print(f"Average Pressure: {p_avg:.2f} bar(a)")
    # print(f"Average Temperature: {t_avg:.2f} C")
    # print(f"Specific Volume: {sf.round(V, 3)} m^3/kg")
    # print(f"Absolute Viscosity: {sf.round(my * 1000, 3)} cP")
    # print(f"Fluid Velocity: {sf.round(vel, 3)} m/s = {sf.round(vel * 60, 3)} m/min")
    # print(f"Reynolds Number: {sf.round(Re,3)}, {turbulence}")
    # print(f"Friction Factor: {sf.round(f,3)}\n")
    
    #Analyzing Pressure Drop

    L_f = D * fitting_resistance
    L_tel = L_f + L_p
    h_l = f * (L_tel / D) * (vel**2 / (2 * g))
    dP = h_l / (10197 * V)
    static_head = static_head_m / (10197 * V)
    misc_pressure_drop = misc_pressure_drop_m / (10197 * V)

    total_pressure_drop = dP + static_head + misc_pressure_drop_m

    # print(f"Sum (nL/D): {fitting_resistance} m\n")
    # print(f"Equivalent Length of Fittings: {L_f}")
    # print(f"Straight Pipe Length: {L_p} m")
    # print(f"Total Equivalent Pipe Length: {L_tel} m")
    # print(f"Friction Head Loss: {h_l:.2f} m")
    # print(f"Friction Pressure Drop: {dP:.2f} bar")
    # print(f"Static Head: {static_head_m} m = {static_head} bar")
    # print(f"Misc. Pressure Drop: {misc_pressure_drop_m} m = {misc_pressure_drop} bar")
    # print("-------------------------")
    # print(f"Total Pressure Drop: {total_pressure_drop:.2f} bar")

    parameters = [D, p_avg, t_avg, V, my, vel, Re, f, turbulence, fitting_resistance, L_f, L_tel, h_l, dP, static_head, misc_pressure_drop]
        
    return total_pressure_drop, parameters

def pipe_to_excel(data_dict, fittings):
    data = pd.DataFrame(data_dict.items(), columns=['Name', 'Value'])

    book = load_workbook('tables/template.xlsx')

    # Check if the sheet exists and remove it
    if 'data' in book.sheetnames:
        del book['data']

    if 'fittings' in book.sheetnames:
        del book['fittings']

    book.save('output_tables/pipe_pressure.xlsx')

    with pd.ExcelWriter('output_tables/pipe_pressure.xlsx', engine='openpyxl', mode='a') as writer:
        data.to_excel(writer, sheet_name='data', index=False)

    with pd.ExcelWriter('output_tables/pipe_pressure.xlsx', engine='openpyxl', mode='a') as writer:
        fittings.to_excel(writer, sheet_name='fittings', index=False)